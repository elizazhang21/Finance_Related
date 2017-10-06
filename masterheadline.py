import openpyxl  # read excel file
#from bs4 import BeautifulSoup
from urllib.request import urlopen
from urllib import request as rq
import requests
import os.path
import datetime
import time
import re
import pandas as pd
import pandas_datareader.data
import numpy as np
from pandas.tseries.offsets import BDay  # time series date function
#import urllib
#import urllib.request as urllib2
import requests
import os
#import ssl
from shutil import copyfile

session = requests.Session()
global count
count = 0


def dateform(sheet):
    ii = 2
    while sheet.cell('A'+str(ii)).value is not None:
        if type(sheet.cell('A'+str(ii)).value) is str:
            sheet.cell('A'+str(ii)).value = datetime.datetime.strptime(
                sheet.cell('A'+str(ii)).value, '%Y-%m-%d')
        ii = ii + 1


def url_open(url):
    print(url)
    status = False
    trytime = 0
    info1 = None
    while not status and trytime <= 5:
        try:
            info1 = urlopen(url, timeout=30)
            encoding = info1.info().get_content_charset('utf-8')
            info1 = info1.read().decode(encoding)
            status = True
        except:
            trytime = trytime + 1
            print('Can not access to the url. Try again after 1 minute.')
            print('It is the ' + str(trytime) + " trial.")
            time.sleep(100)
    return info1


def requests_get(url):
    status = False
    trytime = 0
    info = None
    while not status and trytime <= 5:
        try:
            info = requests.get(url, timeout=30)
            status = True
        except:
            trytime = trytime + 1
            print('Can not access to the url. Try again after 1 minute.')
            print('It is the ' + str(trytime) + " trial.")
            time.sleep(100)
    return info


def GetPriceURL(ticker):
    tickerG = ticker
    if re.search('-', tickerG):
        # ticker is changed if it contains dash
        tickerG = tickerG.replace('-', '.')
    url = 'https://finance.google.com/finance/getprices?i=60&p=3d&f=d,o,h,l,c,v&q='+tickerG
    req = requests.get(url)
    print(req.url)
    html = urlopen(req.url).read().decode()
    try:

        datalist = html.splitlines()
    except:
        print("---- Can't access the page, probably blocked by Google ----")
        datalist = "NULL"
    return datalist

print(GetPriceURL('AAPL'))
