import time
from datetime import datetime as dt
import openpyxl
import pytz
from pytz import timezone


path = ''
filename = 'Street_Account_Data.xlsx'
config_path = ''
config_filename = 'International Market Hours.xlsx'

eastern = timezone('US/Eastern')
eur_1 = timezone('Europe/London')
eur_2 = timezone('Europe/Amsterdam')
eur_3 = timezone('Europe/Athens')
eur_4 = timezone('Europe/Moscow')
aus = timezone('Australia/Sydney')
nzl = timezone('Pacific/Auckland')
ind = timezone('Asia/Kolkata')
chi = timezone('Asia/Shanghai')
jap = timezone('Asia/Tokyo')


tz = {'London': eur_1, 'Portugal': eur_1, 'Irealand': eur_1,
      'France': eur_2, 'Italy': eur_2, 'Germany': eur_2, 'Denmark': eur_2, 'Sweden': eur_2, 'Switzerland': eur_2,
      'Austria': eur_2, 'Netherlands': eur_2, 'Norway': eur_2, 'Belgium': eur_2, 'Spain': eur_2, 'South Africa': eur_2,
      'Greece': eur_3, 'Ukraine': eur_3, 'Finland': eur_3,
      'Russia': eur_4, 'Turkey': eur_4,
      'Canada': eastern, 'Mexico': eastern,
      'Japan': jap, 'Korea': jap,
      'China': chi, 'Hong Kong': chi, 'Singapore': chi,
      'India': ind,
      'Australia': aus, 'New Zealand': nzl,
      }

# ==================================== Load Configuration File ===========
cg = openpyxl.load_workbook(config_path + config_filename)
cg = cg.get_sheet_by_name("Sheet")
Country = []
Open_time = {}
Close_time = {}

for i in range(3, 37):
    country = cg.cell(row=i, column=1).value
    Country.append(country)
    Open_time[country] = cg.cell(row=i, column=4).value
    Close_time[country] = cg.cell(row=i, column=5).value


# ==================================== Read 'Street Account Data' File ====================================
f = openpyxl.load_workbook(path + filename)
sheet = f.get_sheet_by_name("Global Ratings")
n = sheet.max_row + 1
for i in range(2, n):

    ny_date = sheet.cell('A' + str(i)).value
    ny_date = dt.strptime(ny_date, '%m/%d/%y')
    ny_time = sheet.cell('B' + str(i)).value
    ny_time = dt.time(dt.strptime(ny_time, '%H:%M'))
    ny_dt = dt.combine(ny_date, ny_time)
    country = sheet.cell('H' + str(i)).value
    try:
        loc_dt = ny_dt.astimezone(tz[country])
        loc_time = dt.time(loc_dt)
        sheet.cell('D' + str(i)).value = loc_dt.strftime('%m/%d/%y')
        sheet.cell('E' + str(i)).value = loc_dt.strftime('%H:%M')
        if loc_time > Open_time[country] and loc_time < Close_time[country]:
            sheet.cell('AT' + str(i)).value = 'Y'
        else:
            sheet.cell('AT' + str(i)).value = 'N'
    except:
        continue
f.save(filename)
